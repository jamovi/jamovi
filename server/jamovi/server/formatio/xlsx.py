
from datetime import datetime
import math
from numbers import Number

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from jamovi.server.instancemodel import InstanceModel

from .reader import Reader
from .exceptions import FileCorruptError


def get_readers():
    return [ ( 'xlsx', read ) ]


def get_writers():
    return [ ( 'xlsx', write ) ]


def read(data, path, prog_cb, *, settings, **_):

    reader = XLSXReader(settings)
    reader.read_into(data, path, prog_cb)


def write(data: InstanceModel, path, prog_cb):

    wb = Workbook()
    ws = wb.active

    assert ws is not None

    def should_exclude(column):
        return column.is_virtual or (column.is_filter and column.active)

    cols = [ col for col in data if not should_exclude(col) ]
    col_nos = [ col.index for col in cols ]
    col_names = [ col.name for col in cols ]
    col_widths = [ max(len(name), 13) for name in col_names ]

    ws.append(col_names)

    for row_no in range(data.row_count):
        if data.is_row_filtered(row_no):
            continue
        row_values = [None] * len(col_nos)
        for i, col_no in enumerate(col_nos):
            value = data[col_no][row_no]
            row_values[i] = value
            if isinstance(value, Number):
                try:
                    width = int(math.log10(value) // 1)
                except ValueError:
                    width = 1
            else:
                width = len(str(value))
            if width > col_widths[i]:
                col_widths[i] = width
        ws.append(row_values)

        if row_no % 1000 == 0:
            prog_cb(row_no / data.row_count)

    for i in range(len(col_nos)):
        # auto size columns
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = col_widths[i]

    ws.freeze_panes = 'A2'  # freeze first row

    wb.save(path)


def to_string(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        else:
            return str(value)
    else:
        return str(value)


class XLSXReader(Reader):

    _first_col: int = 0
    _last_col: int = 0
    _first_row: int = 0
    _last_row: int = 0
    _row_count: int = 0
    _col_count: int = 0

    def __init__(self, settings):
        Reader.__init__(self, settings)
        self._file = None
        self._ws = None
        self._ws_iter = None
        self._row_no = 0

    def open(self, path):

        self._file = open(path, 'rb')
        wb = load_workbook(self._file, read_only=True, data_only=True)
        self._ws = wb.active

        if self._ws is None:
            raise FileCorruptError

        # survey software often doesn't set these properly
        bad_min_max = (self._ws.max_row is None or self._ws.min_row is None)

        # qualtrics doesn't set these values correctly
        bad_from_qualtrics = 1 == self._ws.min_row == self._ws.min_column == self._ws.max_row == self._ws.max_column

        # xlsx sheets with many rows are typically mostly empty
        # i'm not sure what software is responsible
        many_rows_probably_empty = self._ws.max_row > 500000

        if bad_min_max or bad_from_qualtrics or many_rows_probably_empty:

            self._first_col = 0
            self._last_col = 0
            self._first_row = 0
            self._last_row = 0

            values = self._ws.iter_rows(
                min_row=1,
                min_col=1,
                max_row=1048576,  # apparently openpyxl only reads as many rows as there are
                max_col=1000,
                values_only=True)

            empty_count = 0

            for row_no, row in enumerate(values):
                for col_no, value in enumerate(row):
                    empty_count += 1
                    if value is not None:
                        self._first_col = min(col_no, self._first_col)
                        self._last_col = max(col_no, self._last_col)
                        self._first_row = min(row_no, self._first_row)
                        self._last_row = row_no
                        empty_count = 0

                # if we find 10000 empty rows, probably the rest of the
                # data set is empty
                if empty_count >= 10000:
                    break

        else:
            self._first_row = self._ws.min_row - 1
            self._first_col = self._ws.min_column - 1
            self._last_row = self._ws.max_row - 1
            self._last_col = self._ws.max_column - 1

        self._row_count = self._last_row - self._first_row + 1
        self._col_count = self._last_col - self._first_col + 1

        self.set_total(self._row_count)

    def close(self):
        if self._file is not None:
            self._file.close()
        self._ws = None
        self._ws_iter = None

    def progress(self):
        return self._row_no

    def __iter__(self):
        self._row_no = 0
        self._ws_iter = self._ws.iter_rows(
            min_row=self._first_row + 1,
            min_col=self._first_col + 1,
            max_row=self._last_row + 1,
            max_col=self._last_col + 1,
            values_only=True).__iter__()
        return self

    def __next__(self):
        self._row_no += 1
        values = self._ws_iter.__next__()
        values = list(map(to_string, values))
        return values
